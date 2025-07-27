#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交易数据处理程序
处理交易数据、费率配置和收盘价格，生成盈亏分析、交易明细、持仓数据和分红记录
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import logging

# 导入配置
from config.settings import LOG_CONFIG, SHEET_NAMES, DEFAULT_RATES

# 配置日志
logging.basicConfig(
    level=getattr(logging, LOG_CONFIG['level']),
    format=LOG_CONFIG['format'],
    handlers=[
        logging.FileHandler(LOG_CONFIG['file'], encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('trading_processor')

class TradingProcessor:
    """交易数据处理器类，处理交易数据并生成分析报告"""
    
    def __init__(self):
        """初始化交易数据处理器"""
        self.trades_df = None
        self.rates_df = None
        self.prices_df = None
        self.securities_df = None  # 新增：证券代码信息
        self.dividend_df = None  # 分红记录
        self.fee_rates = {}
        self.positions = {}
        self.daily_pnl = None
    
    def load_data(self, input_file, trades_sheet='交易数据', rates_sheet='费率配置', prices_sheet='收盘价格', securities_sheet='证券信息', dividends_sheet='分红记录'):
        """
        从单个Excel文件的不同工作表加载交易数据、费率配置、收盘价格、证券信息和分红记录
        
        Args:
            input_file: 输入Excel文件路径
            trades_sheet: 交易数据工作表名称
            rates_sheet: 费率配置工作表名称
            prices_sheet: 收盘价格工作表名称
            securities_sheet: 证券信息工作表名称
            dividends_sheet: 分红记录工作表名称
            
        Returns:
            是否成功加载数据
        """
        try:
            # 加载交易数据，指定证券代码为字符串类型
            self.trades_df = pd.read_excel(input_file, sheet_name=trades_sheet, dtype={'证券代码': str})
            logger.info(f"成功从工作表 '{trades_sheet}' 加载 {len(self.trades_df)} 条交易记录")
            
            # 加载费率配置
            self.rates_df = pd.read_excel(input_file, sheet_name=rates_sheet)
            logger.info(f"成功从工作表 '{rates_sheet}' 加载 {len(self.rates_df)} 条费率配置")
            
            # 加载收盘价格，指定证券代码为字符串类型
            self.prices_df = pd.read_excel(input_file, sheet_name=prices_sheet, dtype={'证券代码': str})
            logger.info(f"成功从工作表 '{prices_sheet}' 加载 {len(self.prices_df)} 条收盘价格记录")
            
            # 尝试加载证券信息（可选）
            try:
                self.securities_df = pd.read_excel(input_file, sheet_name=securities_sheet, dtype={'证券代码': str})
                logger.info(f"成功从工作表 '{securities_sheet}' 加载 {len(self.securities_df)} 条证券信息")
            except Exception as e:
                logger.warning(f"未找到证券信息工作表 '{securities_sheet}'，将自动生成证券信息: {e}")
                self.securities_df = None
            
            # 加载分红记录，指定证券代码为字符串类型
            self.dividend_df = pd.read_excel(input_file, sheet_name=dividends_sheet, dtype={'证券代码': str})
            logger.info(f"成功从工作表 '{dividends_sheet}' 加载 {len(self.dividend_df)} 条分红记录")

            # 数据预处理
            self._preprocess_data()
            
            # 处理费率配置
            self._process_fee_rates()
            
            # 处理证券信息
            self._process_securities_info()
            
            return True
        except Exception as e:
            logger.error(f"加载数据失败: {e}")
            return False
    
    def _preprocess_data(self):
        """数据预处理"""
        # 确保日期格式正确
        self.trades_df['日期'] = pd.to_datetime(self.trades_df['日期'])
        self.prices_df['日期'] = pd.to_datetime(self.prices_df['日期'])
        self.dividend_df['日期'] = pd.to_datetime(self.dividend_df['日期'])
        
        # 确保证券代码是字符串类型
        self.trades_df['证券代码'] = self.trades_df['证券代码'].astype(str)
        self.prices_df['证券代码'] = self.prices_df['证券代码'].astype(str)
        self.securities_df['证券代码'] = self.securities_df['证券代码'].astype(str)
        self.dividend_df['证券代码'] = self.dividend_df['证券代码'].astype(str)
        
        # 对数据按照（日期+证券代码）排序
        self.trades_df = self.trades_df.sort_values(['日期', '证券代码']).reset_index(drop=True)
        self.prices_df = self.prices_df.sort_values(['日期', '证券代码']).reset_index(drop=True)
        self.securities_df = self.securities_df.sort_values('证券代码').reset_index(drop=True)
        self.dividend_df = self.dividend_df.sort_values(['日期', '证券代码']).reset_index(drop=True)
        
        # 根据证券代码自动填充交易数据中的证券名称和市场信息
        self._fill_security_info()
    
    def _fill_security_info(self):
        """根据证券代码自动填充交易数据中的证券名称和市场信息"""
        if self.securities_df is None or self.trades_df is None:
            return
        
        logger.info("开始根据证券代码自动填充交易数据中的证券信息")
        
        # 确保交易数据中有证券名称和市场列，如果没有则创建
        if '证券名称' not in self.trades_df.columns:
            self.trades_df['证券名称'] = ''
        if '市场' not in self.trades_df.columns:
            self.trades_df['市场'] = ''
        
        # 为每条交易记录填充证券信息
        for idx, trade in self.trades_df.iterrows():
            symbol = str(trade['证券代码'])
            
            # 在证券信息表中查找对应的信息
            security_info = self.securities_df[self.securities_df['证券代码'] == symbol]
            
            if not security_info.empty:
                # 找到对应的证券信息，填充到交易数据中
                security_name = security_info.iloc[0]['证券名称']
                exchange = security_info.iloc[0]['交易所']
                
                # 根据交易所推断市场
                if exchange == '上交所':
                    market = '上交所'
                elif exchange == '深交所':
                    market = '深交所'
                elif exchange == '港交所':
                    market = '港交所'
                elif exchange == '美股':
                    market = '美股'
                else:
                    market = exchange
                
                # 更新交易数据
                self.trades_df.at[idx, '证券名称'] = security_name
                self.trades_df.at[idx, '市场'] = market
            else:
                # 如果没有找到对应的证券信息，使用默认值或根据代码推断
                logger.warning(f"未找到证券代码 {symbol} 的证券信息，将使用默认值")
                
                # 根据证券代码推断基本信息
                if symbol.startswith(('600', '601', '603', '688', '689', '510', '511', '512', '513', '515', '516', '518')):
                    market = '上交所'
                elif symbol.startswith(('000', '001', '002', '003', '300', '159', '160', '161', '162', '163', '164', '165', '166', '167', '168', '169')):
                    market = '深交所'
                elif len(symbol) == 5 and symbol.isdigit():
                    market = '港交所'
                elif any(c.isalpha() for c in symbol):
                    market = '美股'
                else:
                    market = '未知市场'
                
                # 使用证券代码作为默认名称
                self.trades_df.at[idx, '证券名称'] = symbol
                self.trades_df.at[idx, '市场'] = market
        
        logger.info("证券信息自动填充完成")
        
        # 根据证券代码推断产品类型
        self.trades_df['产品类型'] = self.trades_df['证券代码'].apply(self._infer_product_type)
        logger.info("已根据证券代码推断产品类型")
        
        # 确保收盘价为数值类型，将空值或非数值替换为0
        self.prices_df['收盘价'] = pd.to_numeric(self.prices_df['收盘价'], errors='coerce').fillna(0)
    
    def _process_fee_rates(self):
        """处理费率配置，转换为嵌套字典格式"""
        # 将费率数据转换为嵌套字典格式: {券商: {市场: {产品类型: {费率信息}}}}
        for _, row in self.rates_df.iterrows():
            broker = row['券商']
            market = row['市场']
            product_type = row['产品类型']
            
            if broker not in self.fee_rates:
                self.fee_rates[broker] = {}
            if market not in self.fee_rates[broker]:
                self.fee_rates[broker][market] = {}
            
            self.fee_rates[broker][market][product_type] = {
                '手续费率': row['手续费率'],
                '规费率': row['规费率'],
                '印花税率': row['印花税率'],
                '过户费率': row['过户费率'],
                '最低手续费': row['最低手续费'],
                '平台使用费': row['平台使用费'],
                '结算费': row['结算费'],
                '汇率费': row['汇率费'],
                '监管费': row['监管费']
            }
    
    def _process_securities_info(self):
        """处理证券信息，如果没有提供则自动生成"""
        if self.securities_df is None:
            # 自动生成证券信息
            self._generate_securities_info()
        else:
            # 验证证券信息的完整性
            self._validate_securities_info()
            # 对证券信息进行去重处理
            self._deduplicate_securities_info()
    
    def _deduplicate_securities_info(self):
        """对证券信息进行去重处理"""
        if self.securities_df is None:
            return
        
        # 记录去重前的数量
        original_count = len(self.securities_df)
        
        # 根据证券代码去重，保留第一条记录
        self.securities_df = self.securities_df.drop_duplicates(subset=['证券代码'], keep='first')
        
        # 重新排序和重置索引
        self.securities_df = self.securities_df.sort_values('证券代码').reset_index(drop=True)
        
        # 记录去重后的数量
        deduplicated_count = len(self.securities_df)
        
        if original_count > deduplicated_count:
            logger.info(f"证券信息去重完成：原有 {original_count} 条，去重后 {deduplicated_count} 条，删除了 {original_count - deduplicated_count} 条重复记录")
        else:
            logger.info("证券信息无重复记录")
    
    def update_securities_info_file(self, input_file, securities_sheet='证券信息'):
        """将去重后的证券信息重新写入原文件
        
        Args:
            input_file: 输入Excel文件路径
            securities_sheet: 证券信息工作表名称
        """
        if self.securities_df is None or self.securities_df.empty:
            logger.warning("没有证券信息数据可写入")
            return False
        
        try:
            # 读取原文件的所有工作表
            with pd.ExcelFile(input_file) as xls:
                sheet_names = xls.sheet_names
                all_sheets = {}
                
                # 读取除证券信息外的所有工作表
                for sheet_name in sheet_names:
                    if sheet_name != securities_sheet:
                        all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name, dtype={'证券代码': str})
            
            # 将去重后的证券信息添加到工作表字典中
            all_sheets[securities_sheet] = self.securities_df
            
            # 重新写入文件
            with pd.ExcelWriter(input_file, engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"已将去重后的证券信息重新写入文件: {input_file}")
            return True
            
        except Exception as e:
            logger.error(f"更新证券信息文件失败: {e}")
            return False
    
    def _generate_securities_info(self):
        """根据交易数据自动生成证券信息"""
        if self.trades_df is None:
            logger.warning("没有交易数据，无法生成证券信息")
            return
        
        # 确保证券代码是字符串类型
        self.trades_df['证券代码'] = self.trades_df['证券代码'].astype(str)
        
        # 从交易数据中提取唯一的证券代码和名称
        securities_data = []
        
        # 检查必需的列是否存在
        required_cols = ['证券代码', '证券名称']
        available_cols = list(self.trades_df.columns)
        
        # 确保必需的列存在
        for col in required_cols:
            if col not in available_cols:
                logger.error(f"交易数据中缺少必需的列: {col}")
                return
        
        # 安全地获取列，如果市场列不存在则使用默认值
        if '市场' in available_cols:
            unique_securities = self.trades_df[['证券代码', '证券名称', '市场']].drop_duplicates()
        else:
            unique_securities = self.trades_df[['证券代码', '证券名称']].drop_duplicates()
            unique_securities['市场'] = '默认市场'  # 添加默认市场列
        
        for _, row in unique_securities.iterrows():
            symbol = row['证券代码']
            name = row['证券名称']
            market = row.get('市场', '默认市场')
            
            # 根据市场推断交易所
            if market in ['上交所', '上海','上海证券交易所', 'SSE']:
                exchange = '上交所'
            elif market in ['深交所', '深圳','深圳证券交易所', 'SZSE']:
                exchange = '深交所'
            elif market in ['港交所', '香港', '香港交易所', 'HKEX']:
                exchange = '港交所'
            elif market in ['美股', '美国', 'NASDAQ', 'NYSE']:
                exchange = '美股'
            else:
                # 根据证券代码推断交易所
                if symbol.startswith(('600', '601', '603', '688', '689', '510', '511', '512', '513', '515', '516', '518')):
                    exchange = '上交所'
                elif symbol.startswith(('000', '001', '002', '003', '300', '159', '160', '161', '162', '163', '164', '165', '166', '167', '168', '169')):
                    exchange = '深交所'
                elif len(symbol) == 5 and symbol.isdigit():
                    exchange = '港交所'
                elif any(c.isalpha() for c in symbol):
                    exchange = '美股'
                else:
                    exchange = market  # 使用原始市场信息
            
            securities_data.append({
                '证券代码': symbol,
                '证券名称': name,
                '交易所': exchange
            })
        
        # 创建证券信息DataFrame
        self.securities_df = pd.DataFrame(securities_data)
        self.securities_df = self.securities_df.sort_values('证券代码').reset_index(drop=True)
        
        logger.info(f"已自动生成 {len(self.securities_df)} 条证券信息")
    
    def _validate_securities_info(self):
        """验证证券信息的完整性"""
        if self.securities_df is None or self.trades_df is None:
            return
        
        # 确保证券代码都是字符串类型
        self.trades_df['证券代码'] = self.trades_df['证券代码'].astype(str)
        self.securities_df['证券代码'] = self.securities_df['证券代码'].astype(str)
        
        # 检查交易数据中的证券是否都在证券信息中
        trade_symbols = set(self.trades_df['证券代码'].unique())
        securities_symbols = set(self.securities_df['证券代码'].unique())
        
        missing_symbols = trade_symbols - securities_symbols
        if missing_symbols:
            logger.warning(f"证券信息中缺少以下证券代码: {missing_symbols}")
            
            # 自动补充缺失的证券信息
            for symbol in missing_symbols:
                trade_info = self.trades_df[self.trades_df['证券代码'] == symbol].iloc[0]
                name = trade_info['证券名称']
                market = trade_info.get('市场', '默认市场')
                
                # 推断交易所
                if market in ['上交所', '上海证券交易所', 'SSE']:
                    exchange = '上交所'
                elif market in ['深交所', '深圳证券交易所', 'SZSE']:
                    exchange = '深交所'
                elif market in ['港交所', '香港交易所', 'HKEX']:
                    exchange = '港交所'
                elif market in ['美股', 'NASDAQ', 'NYSE']:
                    exchange = '美股'
                else:
                    exchange = market
                
                # 添加到证券信息中
                new_row = pd.DataFrame({
                    '证券代码': [symbol],
                    '证券名称': [name],
                    '交易所': [exchange]
                })
                self.securities_df = pd.concat([self.securities_df, new_row], ignore_index=True)
            
            # 重新排序
            self.securities_df = self.securities_df.sort_values('证券代码').reset_index(drop=True)
            logger.info(f"已补充 {len(missing_symbols)} 条缺失的证券信息")
        
        logger.info("证券信息验证完成")
    
    def get_security_info(self, symbol):
        """根据证券代码获取证券信息
        
        Args:
            symbol: 证券代码
            
        Returns:
            包含证券信息的字典，如果未找到则返回默认值
        """
        if self.securities_df is None:
            return {'证券名称': '', '交易所': ''}
        
        # 确保证券代码是字符串类型
        symbol = str(symbol)
        
        # 在证券信息表中查找
        security_info = self.securities_df[self.securities_df['证券代码'] == symbol]
        
        if not security_info.empty:
            return {
                '证券名称': security_info.iloc[0]['证券名称'],
                '交易所': security_info.iloc[0]['交易所']
            }
        else:
            return {'证券名称': '', '交易所': ''}
        
        # 检查交易数据中的证券是否都在证券信息中
        trade_symbols = set(self.trades_df['证券代码'].unique())
        securities_symbols = set(self.securities_df['证券代码'].unique())
        
        missing_symbols = trade_symbols - securities_symbols
        if missing_symbols:
            logger.warning(f"证券信息中缺少以下证券代码: {missing_symbols}")
            
            # 自动补充缺失的证券信息
            for symbol in missing_symbols:
                trade_info = self.trades_df[self.trades_df['证券代码'] == symbol].iloc[0]
                name = trade_info['证券名称']
                market = trade_info.get('市场', '默认市场')
                
                # 推断交易所
                if market in ['上交所', '上海证券交易所', 'SSE']:
                    exchange = '上交所'
                elif market in ['深交所', '深圳证券交易所', 'SZSE']:
                    exchange = '深交所'
                elif market in ['港交所', '香港交易所', 'HKEX']:
                    exchange = '港交所'
                elif market in ['美股', 'NASDAQ', 'NYSE']:
                    exchange = '美股'
                else:
                    exchange = market
                
                # 添加到证券信息中
                new_row = pd.DataFrame({
                    '证券代码': [symbol],
                    '证券名称': [name],
                    '交易所': [exchange]
                })
                self.securities_df = pd.concat([self.securities_df, new_row], ignore_index=True)
            
            # 重新排序
            self.securities_df = self.securities_df.sort_values('证券代码').reset_index(drop=True)
            logger.info(f"已补充 {len(missing_symbols)} 条缺失的证券信息")
        
        logger.info("证券信息验证完成")
    
    def _infer_product_type(self, code):
        """根据证券代码推断产品类型"""
        code = str(code).upper()
        
        # A股股票
        if code.startswith(('000', '001', '002', '003', '300', '600', '601', '603', '688', '689')):
            return '股票'
        # ETF基金
        elif code.startswith(('159', '510', '511', '512', '513', '515', '516', '518')):
            return 'ETF'
        # 场内基金
        elif code.startswith(('160', '161', '162', '163', '164', '165', '166', '167', '168', '169')):
            return '基金'
        # 港股
        elif len(code) == 5 and code.isdigit():
            return '港股'
        # 美股（简单判断，包含字母）
        elif any(c.isalpha() for c in code):
            return '美股'
        else:
            return '股票'  # 默认为股票
    
    def calculate_fees(self):
        """计算交易费用"""
        if self.trades_df is None:
            logger.error("请先加载交易数据")
            return False
        
        # 计算交易金额
        self.trades_df['交易金额'] = self.trades_df['成交价格'] * self.trades_df['成交数量']
        
        # 初始化费用列
        fee_columns = ['手续费', '规费', '印花税', '过户费', '平台使用费', '结算费', '汇率费', '监管费', '总费用']
        for col in fee_columns:
            self.trades_df[col] = 0.0
        
        # 计算每笔交易的费用
        for idx, trade in self.trades_df.iterrows():
            # 安全获取字段值，如果不存在则使用默认值
            broker = trade.get('券商', '默认券商')
            market = trade.get('市场', '默认市场')
            product_type = trade.get('产品类型', '股票')
            amount = trade['交易金额']
            is_sell = trade['买卖方向'] in ['卖出', '卖', 'SELL', 'S']
            
            # 获取费率设置
            if (broker in self.fee_rates and 
                market in self.fee_rates[broker] and 
                product_type in self.fee_rates[broker][market]):
                rates = self.fee_rates[broker][market][product_type]
            else:
                logger.warning(f"未找到券商 {broker} 市场 {market} 产品类型 {product_type} 的费率设置，使用默认费率")
                rates = {
                    '手续费率': 0.0003,
                    '规费率': 0,
                    '印花税率': 0.001 if product_type == '股票' else 0,
                    '过户费率': 0.00002 if product_type == '股票' else 0,
                    '最低手续费': 5,
                    '平台使用费': 0,
                    '结算费': 0,
                    '汇率费': 0,
                    '监管费': 0
                }
            
            # 计算各项费用
            if broker == '国泰君安':            # 这里的逻辑是根据君安app测算出来的
                commission = amount * rates['手续费率']
                gui_fee = amount * rates['规费率']

                # 如果低于最低手续费，那么手续费+规费是5（君安的规则）
                if commission < rates['最低手续费']:
                    commission = rates['最低手续费'] - gui_fee
            else:
                commission = max(amount * rates['手续费率'], rates['最低手续费'])
                gui_fee = amount * rates['规费率']
            
            stamp_tax = amount * rates['印花税率'] if is_sell else 0
            transfer_fee = amount * rates['过户费率']
            platform_fee = rates['平台使用费']
            settlement_fee = amount * rates['结算费']
            fx_fee = amount * rates['汇率费'] if market in ['港交所', '美股'] else 0
            regulatory_fee = amount * rates['监管费']
            
            total_fee = commission + stamp_tax + transfer_fee + platform_fee + settlement_fee + fx_fee + regulatory_fee+gui_fee
            
            # 更新交易记录中的费用，保留2位小数
            self.trades_df.at[idx, '手续费'] = round(commission, 2)
            self.trades_df.at[idx, '规费'] = round(gui_fee, 2)
            self.trades_df.at[idx, '印花税'] = round(stamp_tax, 2)
            self.trades_df.at[idx, '过户费'] = round(transfer_fee, 2)
            self.trades_df.at[idx, '平台使用费'] = round(platform_fee, 2)
            self.trades_df.at[idx, '结算费'] = round(settlement_fee, 2)
            self.trades_df.at[idx, '汇率费'] = round(fx_fee, 2)
            self.trades_df.at[idx, '监管费'] = round(regulatory_fee, 2)
            self.trades_df.at[idx, '总费用'] = round(total_fee, 2)
        
        logger.info("交易费用计算完成")
        return True
    
    def update_positions(self):
        """根据交易记录更新持仓情况"""
        if self.trades_df is None:
            logger.error("请先加载交易数据")
            return False
        
        # 确保交易记录按日期排序
        self.trades_df = self.trades_df.sort_values('日期')
        
        # 遍历每笔交易，更新持仓
        for _, trade in self.trades_df.iterrows():
            symbol = trade['证券代码']
            date = trade['日期']
            price = trade['成交价格']
            quantity = trade['成交数量']
            is_buy = trade['买卖方向'] not in ['卖出', '卖', 'SELL', 'S']
            
            # 更新持仓
            if symbol not in self.positions:
                # 新建持仓
                if is_buy:
                    self.positions[symbol] = {
                        '证券名称': trade['证券名称'],
                        '持仓数量': quantity,
                        '持仓成本': price,
                        '市场': trade.get('市场', '默认市场'),
                        '产品类型': trade.get('产品类型', '股票'),
                        '每日价格': {date.strftime('%Y-%m-%d'): price}
                    }
            else:
                # 更新现有持仓
                position = self.positions[symbol]
                current_qty = position['持仓数量']
                current_cost = position['持仓成本']
                
                if is_buy:
                    # 买入，增加持仓
                    new_qty = current_qty + quantity
                    # 更新加权平均成本（包含手续费）
                    fees = trade.get('总费用', 0)
                    new_cost = (current_qty * current_cost + quantity * price + fees) / new_qty if new_qty > 0 else 0
                    position['持仓数量'] = new_qty
                    position['持仓成本'] = new_cost
                else:
                    # 卖出，先用卖出前的成本价计算已实现盈亏
                    # 注意：这里只计算盈亏，但不记录，因为这个方法只负责更新持仓
                    realized_pnl = (price - current_cost) * quantity - trade.get('总费用', 0)
                    
                    # 然后再减少持仓
                    new_qty = current_qty - quantity
                    position['持仓数量'] = new_qty
                    
                    # 如果持仓数量变为0，成本也应该为0
                    if new_qty <= 0:
                        position['持仓成本'] = 0
                    # 否则保持原有成本价不变
                    # 注意：这里不调整成本价，因为这个方法只负责简单的持仓更新
                
                # 记录每日价格
                position['每日价格'][date.strftime('%Y-%m-%d')] = price
        
        logger.info(f"持仓更新完成，共 {len(self.positions)} 只证券")
        return True

    def get_current_positions(self):
        """获取当前持仓数据 - 每支股票的最新持仓汇总"""
        if self.daily_pnl is None or self.daily_pnl.empty:
            logger.warning("没有持仓数据")
            return pd.DataFrame()
        
        # 按证券代码分组，获取每支股票的最新持仓数据
        positions_data = []
        
        # 获取每支股票的最新记录
        for symbol in self.daily_pnl['证券代码'].unique():
            symbol_data = self.daily_pnl[self.daily_pnl['证券代码'] == symbol]
            
            if symbol_data.empty:
                continue
            
            # 获取最新的记录（按日期排序后的最后一条）
            if '日期' in symbol_data.columns:
                latest_record = symbol_data.sort_values('日期').iloc[-1]
            else:
                latest_record = symbol_data.iloc[-1]
            
            # 检查列名，使用'持仓数量'或'当前持仓数量'
            position_qty_col = '持仓数量' if '持仓数量' in latest_record else '当前持仓数量'
            
            # 只保存有持仓的证券
            if position_qty_col in latest_record and latest_record[position_qty_col] > 0:
                # 检查成本价列名
                if '持仓成本价' in latest_record:
                    cost_price = latest_record['持仓成本价']
                elif '移动平均成本' in latest_record:
                    cost_price = latest_record['移动平均成本']
                elif '当前成本价' in latest_record:
                    cost_price = latest_record['当前成本价']
                else:
                    cost_price = 0
                
                # 检查价格列名
                if '当前价格' in latest_record:
                    price = latest_record['当前价格']
                elif '收盘价' in latest_record:
                    price = latest_record['收盘价']
                else:
                    price = 0
                
                # 检查市值列名
                if '持仓市值' in latest_record:
                    market_value = latest_record['持仓市值']
                elif '当前市值' in latest_record:
                    market_value = latest_record['当前市值']
                else:
                    market_value = latest_record[position_qty_col] * price
                
                # 检查盈亏列名
                realized_pnl = 0
                if '累计已实现盈亏' in latest_record:
                    realized_pnl = latest_record['累计已实现盈亏']
                elif '已实现盈亏' in latest_record:
                    realized_pnl = latest_record['已实现盈亏']
                
                unrealized_pnl = 0
                if '当日未实现盈亏' in latest_record:
                    unrealized_pnl = latest_record['当日未实现盈亏']
                elif '未实现盈亏' in latest_record:
                    unrealized_pnl = latest_record['未实现盈亏']
                
                total_pnl = 0
                if '累计总盈亏' in latest_record:
                    total_pnl = latest_record['累计总盈亏']
                elif '总盈亏' in latest_record:
                    total_pnl = latest_record['总盈亏']
                
                pnl_ratio = 0
                if '总盈亏比例(%)' in latest_record:
                    pnl_ratio = latest_record['总盈亏比例(%)']
                elif '盈亏比例(%)' in latest_record:
                    pnl_ratio = latest_record['盈亏比例(%)']
                
                # 计算交易统计信息
                symbol_trades = self.trades_df[self.trades_df['证券代码'] == symbol]
                buy_trades = symbol_trades[~symbol_trades['买卖方向'].isin(['卖出', '卖', 'SELL', 'S'])]
                sell_trades = symbol_trades[symbol_trades['买卖方向'].isin(['卖出', '卖', 'SELL', 'S'])]
                
                # 买入统计
                total_buy_volume = buy_trades['成交数量'].sum()
                total_buy_amount = (buy_trades['成交价格'] * buy_trades['成交数量']).sum()
                avg_buy_price = total_buy_amount / total_buy_volume if total_buy_volume > 0 else 0
                total_buy_fees = buy_trades['总费用'].sum()
                
                # 卖出统计
                total_sell_volume = sell_trades['成交数量'].sum()
                total_sell_amount = (sell_trades['成交价格'] * sell_trades['成交数量']).sum()
                avg_sell_price = total_sell_amount / total_sell_volume if total_sell_volume > 0 else 0
                total_sell_fees = sell_trades['总费用'].sum()
                
                # 交易次数和时间跨度
                trade_count = len(symbol_trades)
                if not symbol_trades.empty:
                    first_trade_date = symbol_trades['日期'].min().date()
                    last_trade_date = symbol_trades['日期'].max().date()
                    holding_days = (last_trade_date - first_trade_date).days + 1
                else:
                    first_trade_date = None
                    last_trade_date = None
                    holding_days = 0
                
                positions_data.append({
                    '证券代码': latest_record['证券代码'],
                    '证券名称': latest_record['证券名称'],
                    '交易所': latest_record['交易所'],
                    '持仓数量': latest_record[position_qty_col],
                    '持仓成本价': round(cost_price, 4),
                    '当前价格': round(price, 4),
                    '持仓市值': round(market_value, 2),
                    '持仓成本总额': round(latest_record['持仓成本总额'], 2),
                    '已实现盈亏': round(realized_pnl, 2),
                    '未实现盈亏': round(unrealized_pnl, 2),
                    '总盈亏': round(total_pnl, 2),
                    '买入手续费': round(total_buy_fees, 2),
                    '卖出手续费': round(total_sell_fees, 2),
                    '总手续费': round(total_buy_fees + total_sell_fees, 2),
                    '平均买入价': round(avg_buy_price, 4),
                    '平均卖出价': round(avg_sell_price, 4),
                    '交易次数': trade_count,
                    '首次交易日期': first_trade_date.strftime('%Y-%m-%d') if first_trade_date else '',
                    '最后交易日期': last_trade_date.strftime('%Y-%m-%d') if last_trade_date else '',
                    '持有天数': holding_days
                })
        
        # 按总盈亏从大到小排序
        positions_df = pd.DataFrame(positions_data)
        if not positions_df.empty:
            positions_df = positions_df.sort_values('持仓数量', ascending=False)
        
        logger.info(f"当前持仓数据生成完成，共 {len(positions_df)} 只股票")
        return positions_df
    
    def get_stock_historical_pnl(self):
        """
        获取每支股票的历史盈亏数据，按盈亏额从大到小排列
        
        使用核心盈亏计算方法，确保与其他盈亏计算保持一致
        
        Returns:
            DataFrame: 股票历史盈亏数据
        """
        try:
            # 首先确保已经计算了每日盈亏
            if self.daily_pnl is None:
                # 调用核心盈亏计算方法
                daily_positions, pnl_data, _ = self.calculate_pnl_core()
                if pnl_data is None:
                    logger.warning("盈亏核心计算失败，无法生成股票历史盈亏")
                    return pd.DataFrame()
                
                # 创建每日盈亏DataFrame
                self.daily_pnl = pd.DataFrame(pnl_data)
                
                # 按日期和证券代码排序
                self.daily_pnl = self.daily_pnl.sort_values(['日期', '证券代码'])
            
            # 如果每日盈亏数据为空，返回空DataFrame
            if self.daily_pnl.empty:
                logger.warning("没有每日盈亏数据，无法生成股票历史盈亏")
                return pd.DataFrame()
            
            stock_pnl_data = []
            
            # 按证券代码分组统计
            for symbol in self.daily_pnl['证券代码'].unique():
                # 获取该证券的所有记录，按日期排序
                symbol_data = self.daily_pnl[self.daily_pnl['证券代码'] == symbol].sort_values('日期')
                
                if symbol_data.empty:
                    continue
                
                # 获取最新记录（最后一条记录）
                latest_record = symbol_data.iloc[-1]
                
                # 计算交易统计
                symbol_trades = self.trades_df[self.trades_df['证券代码'] == symbol]
                buy_trades = symbol_trades[~symbol_trades['买卖方向'].isin(['卖出', '卖', 'SELL', 'S'])]
                sell_trades = symbol_trades[symbol_trades['买卖方向'].isin(['卖出', '卖', 'SELL', 'S'])]
                
                # 买入统计
                total_buy_volume = buy_trades['成交数量'].sum()
                total_buy_amount = (buy_trades['成交价格'] * buy_trades['成交数量']).sum()
                avg_buy_price = total_buy_amount / total_buy_volume if total_buy_volume > 0 else 0
                total_buy_fees = buy_trades['总费用'].sum()
                
                # 卖出统计
                total_sell_volume = sell_trades['成交数量'].sum()
                total_sell_amount = (sell_trades['成交价格'] * sell_trades['成交数量']).sum()
                avg_sell_price = total_sell_amount / total_sell_volume if total_sell_volume > 0 else 0
                total_sell_fees = sell_trades['总费用'].sum()
                
                # 从最新记录中获取盈亏数据，确保与每日盈亏计算保持一致
                current_volume = latest_record['持仓数量']
                current_cost_price = latest_record['持仓成本价']
                current_price = latest_record['收盘价']
                current_market_value = latest_record['持仓市值']
                cumulative_realized_pnl = latest_record['累计已实现盈亏']
                unrealized_pnl = latest_record['当日未实现盈亏']
                total_pnl = latest_record['总盈亏']
                
                # 计算盈亏比例
                cost_total = latest_record['持仓成本总额']
                pnl_ratio = (total_pnl / cost_total * 100) if cost_total > 0 else 0
                
                # 交易次数和时间跨度
                trade_count = len(symbol_trades)
                if not symbol_trades.empty:
                    first_trade_date = symbol_trades['日期'].min().date()
                    last_trade_date = symbol_trades['日期'].max().date()
                    holding_days = (last_trade_date - first_trade_date).days + 1
                else:
                    first_trade_date = None
                    last_trade_date = None
                    holding_days = 0
                    
                # 添加到股票历史盈亏数据
                stock_pnl_data.append({
                    '证券代码': symbol,
                    '证券名称': latest_record['证券名称'],
                    '交易所': latest_record['交易所'],
                    '累计买入数量': total_buy_volume,
                    '累计买入金额': round(total_buy_amount, 2),
                    '平均买入价': round(avg_buy_price, 4),
                    '累计卖出数量': total_sell_volume,
                    '累计卖出金额': round(total_sell_amount, 2),
                    '平均卖出价': round(avg_sell_price, 4),
                    '当前持仓数量': current_volume,
                    '当前价格': round(current_price, 4),
                    '当前成本价': round(current_cost_price, 4),
                    '当前市值': round(current_market_value, 2),
                    '已实现盈亏': round(cumulative_realized_pnl, 2),
                    '未实现盈亏': round(unrealized_pnl, 2),
                    '总盈亏': round(total_pnl, 2),
                    '盈亏比例(%)': round(pnl_ratio, 2),
                    '交易次数': trade_count,
                    '买入手续费': round(total_buy_fees, 2),
                    '卖出手续费': round(total_sell_fees, 2),
                    '总手续费': round(total_buy_fees + total_sell_fees, 2),
                    '首次交易日期': first_trade_date.strftime('%Y-%m-%d') if first_trade_date else '',
                    '最后交易日期': last_trade_date.strftime('%Y-%m-%d') if last_trade_date else '',
                    '持有天数': holding_days
                })
        
            # 转换为DataFrame并按总盈亏从大到小排序
            stock_pnl_df = pd.DataFrame(stock_pnl_data)
            if not stock_pnl_df.empty:
                stock_pnl_df = stock_pnl_df.sort_values('总盈亏', ascending=False)
            
            logger.info(f"股票历史盈亏数据生成完成，共 {len(stock_pnl_df)} 只股票")
            return stock_pnl_df
        except Exception as e:
            logger.error(f"生成股票历史盈亏数据失败: {e}")
            return pd.DataFrame()
    
    def calculate_pnl_core(self):
        """
        核心盈亏计算方法，使用统一的摊薄成本法
        
        计算方法：
        1. 按日期和证券代码处理交易
        2. 使用摊薄成本法计算每日已实现盈亏和未实现盈亏
        3. 记录每日持仓和盈亏数据
        
        Returns:
            tuple: (daily_positions, daily_pnl_data, all_dates)
            - daily_positions: 每个证券每天的持仓情况 {证券代码: {日期: {'持仓数量': 数量, '持仓成本': 成本价, ...}}}
            - daily_pnl_data: 每日盈亏数据列表
            - all_dates: 所有交易和价格日期的有序列表
        """
        if self.trades_df is None:
            logger.error("请先加载交易数据")
            return None, None, None
            
        # 获取所有交易日期和价格日期
        trade_dates = set(self.trades_df['日期'].dt.date)
        price_dates = set(self.prices_df['日期'].dt.date)
        all_dates = sorted(trade_dates.union(price_dates))
        
        # 创建每日盈亏数据列表
        pnl_data = []
        
        # 记录每个证券每天的持仓情况
        # 格式: {证券代码: {日期: {'持仓数量': 数量, '持仓成本': 成本价, '持仓成本总额': 成本总额, ...}}}
        daily_positions = {}
        
        # 记录已实现盈亏
        # 格式: {证券代码: {日期: 已实现盈亏}}
        realized_pnl = {}
        
        # 初始化字典
        for symbol in set(self.trades_df['证券代码']):
            daily_positions[symbol] = {}
            realized_pnl[symbol] = {}
            for date in all_dates:
                realized_pnl[symbol][date] = 0
        
        # 按日期处理交易，更新持仓
        for date in all_dates:
            # 获取当日所有交易
            day_trades = self.trades_df[self.trades_df['日期'].dt.date == date]
            
            # 处理每个证券
            for symbol in set(self.trades_df['证券代码']):
                # 获取前一天的持仓情况
                prev_dates = [d for d in daily_positions[symbol].keys() if d < date]
                if prev_dates:
                    prev_date = max(prev_dates)
                    prev_position = daily_positions[symbol][prev_date].copy()
                else:
                    # 如果没有前一天的持仓记录，初始化为0
                    prev_position = {
                        '持仓数量': 0,
                        '持仓成本': 0,
                        '持仓成本总额': 0,
                        '证券名称': '',
                        '市场': '',
                        '产品类型': '',
                        '累计已实现盈亏': 0  # 添加累计已实现盈亏字段
                    }
                
                # 初始化当日持仓为前一天的持仓
                current_position = prev_position.copy()
                
                # 当日已实现盈亏
                day_realized_pnl = 0
                
                # 处理当日该证券的所有交易
                symbol_trades = day_trades[day_trades['证券代码'] == symbol]
                for _, trade in symbol_trades.iterrows():
                    price = trade['成交价格']
                    quantity = trade['成交数量']
                    is_buy = trade['买卖方向'] not in ['卖出', '卖', 'SELL', 'S']
                    fees = trade['总费用']
                    
                    # 更新证券基本信息
                    current_position['证券名称'] = trade['证券名称']
                    current_position['市场'] = trade.get('市场', '默认市场')
                    current_position['产品类型'] = trade.get('产品类型', '股票')
                    
                    if is_buy:
                        # 买入，增加持仓
                        old_qty = current_position['持仓数量']
                        old_cost_total = current_position['持仓成本总额']
                        
                        # 新增成本 = 买入金额 + 手续费
                        new_cost_total = price * quantity + fees
                        
                        # 更新持仓数量和成本总额
                        current_position['持仓数量'] = old_qty + quantity
                        current_position['持仓成本总额'] = old_cost_total + new_cost_total
                        
                        # 更新平均成本价
                        if current_position['持仓数量'] > 0:
                            current_position['持仓成本'] = current_position['持仓成本总额'] / current_position['持仓数量']
                        else:
                            current_position['持仓成本'] = 0
                    else:
                        # 卖出，减少持仓并计算已实现盈亏
                        old_qty = current_position['持仓数量']
                        old_cost_total = current_position['持仓成本总额']
                        old_cost_price = current_position['持仓成本']
                        
                        if old_qty > 0:
                            # 计算卖出部分的成本
                            sell_cost_ratio = min(quantity / old_qty, 1.0)  # 确保比例不超过1
                            sell_cost = old_cost_total * sell_cost_ratio
                            
                            # 计算卖出金额（减去手续费）
                            sell_amount = price * quantity - fees
                            
                            # 计算已实现盈亏 = 卖出金额 - 卖出部分的成本
                            trade_realized_pnl = sell_amount - sell_cost
                            
                            # 累加到当日已实现盈亏
                            day_realized_pnl += trade_realized_pnl
                            
                            # 更新持仓
                            current_position['持仓数量'] = old_qty - quantity  
                            
                            # 更新成本总额（减去卖出部分的成本）
                            if current_position['持仓数量'] > 0:
                                # 卖出部分持仓时，成本总额按比例减少，但成本价保持不变
                                current_position['持仓成本总额'] = old_cost_total * (1 - sell_cost_ratio)
                                # 卖出时成本价不变
                            else:
                                # 全部卖出，清零成本
                                current_position['持仓成本'] = 0
                                current_position['持仓成本总额'] = 0
                
                # 记录当日已实现盈亏
                realized_pnl[symbol][date] = day_realized_pnl
                
                # 更新累计已实现盈亏
                current_position['累计已实现盈亏'] = current_position.get('累计已实现盈亏', 0) + day_realized_pnl
                
                # 更新当日持仓情况
                daily_positions[symbol][date] = current_position
                
                # 获取当日收盘价
                price_rows = self.prices_df[(self.prices_df['日期'].dt.date == date) & 
                                           (self.prices_df['证券代码'] == symbol)]
                
                close_price = 0
                if not price_rows.empty and price_rows['收盘价'].values[0] > 0:
                    close_price = price_rows['收盘价'].values[0]
                else:
                    # 如果没有收盘价或收盘价为0，尝试使用当日交易价格
                    if not symbol_trades.empty:
                        # 使用当日最后一笔交易的价格
                        close_price = symbol_trades.iloc[-1]['成交价格']
                
                # 计算未实现盈亏
                qty = current_position['持仓数量']
                cost_price = current_position['持仓成本']
                cost_total = current_position['持仓成本总额']
                market_value = qty * close_price if close_price > 0 else 0
                
                # 统一计算公式：未实现盈亏 = (当前收盘价 - 当前持仓成本价) × 当前持仓数量
                unrealized_pnl = qty * (close_price - cost_price) if qty > 0 else 0
                
                # 计算盈亏比例
                unrealized_pnl_ratio = (unrealized_pnl / cost_total * 100) if cost_total > 0 else 0
                
                # 计算总盈亏 = 累计已实现盈亏 + 未实现盈亏
                total_pnl = current_position['累计已实现盈亏'] + unrealized_pnl
                
                # 添加到盈亏数据
                if qty > 0 or day_realized_pnl != 0:  # 只记录有持仓或有盈亏的日期
                    # 获取证券信息
                    security_info = self.get_security_info(symbol)
                    
                    pnl_data.append({
                        '日期': date,
                        '证券代码': symbol,
                        '证券名称': current_position['证券名称'],
                        '交易所': security_info['交易所'],
                        '持仓数量': qty,
                        '持仓成本价': round(cost_price, 4),
                        '持仓成本总额': round(cost_total, 2),
                        '收盘价': round(close_price, 4),
                        '持仓市值': round(market_value, 2),
                        '当日已实现盈亏': round(day_realized_pnl, 2),
                        '累计已实现盈亏': round(current_position['累计已实现盈亏'], 2),
                        '当日未实现盈亏': round(unrealized_pnl, 2),
                        '未实现盈亏比例(%)': round(unrealized_pnl_ratio, 2),
                        '总盈亏': round(total_pnl, 2)
                    })
        
        # 更新最终持仓到 self.positions
        for symbol, dates in daily_positions.items():
            if dates:
                last_date = max(dates.keys())
                last_position = dates[last_date]
                
                if last_position['持仓数量'] > 0:
                    self.positions[symbol] = {
                        '证券名称': last_position['证券名称'],
                        '持仓数量': last_position['持仓数量'],
                        '持仓成本': last_position['持仓成本'],
                        '市场': last_position['市场'],
                        '产品类型': last_position['产品类型'],
                        '每日价格': {}
                    }
        
        return daily_positions, pnl_data, all_dates
    
    def calculate_daily_pnl(self):
        """
        计算每日盈亏，使用统一的摊薄成本法
        
        计算方法：
        1. 调用核心盈亏计算方法
        2. 创建每日盈亏DataFrame
        
        Returns:
            bool: 是否成功计算盈亏
        """
        try:
            # 调用核心盈亏计算方法
            _, pnl_data, _ = self.calculate_pnl_core()
            
            if pnl_data is None:
                return False
            
            # 创建盈亏DataFrame
            self.daily_pnl = pd.DataFrame(pnl_data)
            
            # 按日期和证券代码排序
            self.daily_pnl = self.daily_pnl.sort_values(['日期', '证券代码'])
            
            logger.info("每日盈亏计算完成，使用摊薄成本法")
            return True
        except Exception as e:
            logger.error(f"计算每日盈亏失败: {e}")
            return False
    
    def process_data(self):
        """处理数据并生成分析结果"""
        # 计算交易费用
        if not self.calculate_fees():
            return False
        
        # 更新持仓情况
        if not self.update_positions():
            return False
        
        # 将持仓数据同步到日志
        logger.info("持仓成本价格情况:")
        for symbol, position in self.positions.items():
            if position['持仓数量'] > 0:
                logger.info(f"证券代码: {symbol}, 持仓数量: {position['持仓数量']}, 持仓成本价: {position['持仓成本']:.4f}")
        
        # 使用摊薄成本法计算每日盈亏
        try:
            if not self.calculate_daily_pnl():
                logger.warning("每日盈亏计算失败")
                return False
            logger.info("每日盈亏计算成功")
        except Exception as e:
            logger.error(f"每日盈亏计算出错: {e}")
            return False
        
        return True
    
    def _format_sheet(self, writer, sheet_name, sheet_type='default'):
        """统一格式化工作表，美化输出
        
        Args:
            writer: ExcelWriter对象
            sheet_name: 工作表名称
            sheet_type: 工作表类型，用于应用特定格式
        """
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 获取工作表
            worksheet = writer.sheets[sheet_name]
            
            # 定义统一样式
            # 统一的表头样式 - 深蓝色背景，白色粗体字
            header_fill = PatternFill(start_color='7586C2', end_color='7586C2', fill_type='solid')      #深蓝色 7586C2
            header_font = Font(color='FFFFFF', bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            right_alignment = Alignment(horizontal='right', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # 定义边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 定义数据行样式
            positive_font = Font(color='006600', bold=True)  # 绿色（盈利）
            negative_font = Font(color='CC0000', bold=True)  # 红色（亏损）
            money_font = Font(color='006600', bold=True)     # 绿色（金额）
            date_font = Font(color='000000')                 # 普通字体（日期）
            
            # 定义盈亏特定样式
            positive_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 浅绿色背景
            negative_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # 浅红色背景

            # 定义交易类型特定样式
            buy_fill = PatternFill(start_color='FBECDE', end_color='FBECDE', fill_type='solid')  # 浅红色 F6C2CA
            sell_fill = PatternFill(start_color='DEEFE0', end_color='DEEFE0', fill_type='solid')  # 浅绿色 DEEFE0
            
            
            # 格式化表头行
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'
            
            # 查找特定列
            column_indices = {}
            for col_idx, cell in enumerate(worksheet[1], 1):
                col_name = str(cell.value)
                column_indices[col_name] = col_idx
            
            # 应用特定格式
            if sheet_type == 'trades':
                # 交易明细特定格式
                direction_col = column_indices.get('买卖方向')
                if direction_col:
                    for row_idx in range(2, worksheet.max_row + 1):
                        direction_cell = worksheet.cell(row=row_idx, column=direction_col)
                        direction_value = str(direction_cell.value).strip()
                        
                        # 为每个单元格添加边框
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                        
                        # 判断买卖方向并应用格式
                        if direction_value in ['买入', '买', 'BUY', 'B']:
                            # 买入：红色背景和字体
                            for col_idx in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = buy_fill
                                if col_idx == direction_col:
                                    cell.font = Font(color='CC0000', bold=True)
                        elif direction_value in ['卖出', '卖', 'SELL', 'S']:
                            # 卖出：绿色背景和字体
                            for col_idx in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = sell_fill
                                if col_idx == direction_col:
                                    cell.font = Font(color='006600', bold=True)
                
                # 格式化金额列
                money_cols = ['成交价格', '成交数量', '交易金额', '手续费', '规费', '印花税', '过户费', '总费用']
                for col_name in money_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = right_alignment
                            if '价格' in col_name:
                                cell.number_format = '#,##0.0000'
                            else:
                                cell.number_format = '#,##0.00'
                
                # 格式化日期列
                if '日期' in column_indices:
                    col_idx = column_indices['日期']
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_alignment
                        cell.number_format = 'yyyy-mm-dd'
                        cell.font = date_font
            
            elif sheet_type in ['pnl', 'stock_pnl']:
                # 盈亏分析特定格式
                total_pnl_col = column_indices.get('总盈亏')
                
                # 找到盈亏相关列
                pnl_columns = {}
                for col_name, col_idx in column_indices.items():
                    if '盈亏' in col_name or '比例' in col_name:
                        pnl_columns[col_name] = col_idx
                
                # 格式化数据行
                for row_idx in range(2, worksheet.max_row + 1):
                    # 为每个单元格添加边框
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                    
                    # 根据总盈亏决定整行背景色
                    if total_pnl_col:
                        total_pnl_cell = worksheet.cell(row=row_idx, column=total_pnl_col)
                        try:
                            total_pnl_value = float(total_pnl_cell.value) if total_pnl_cell.value is not None else 0
                            row_fill = positive_fill if total_pnl_value > 0 else negative_fill
                            
                            # 应用整行背景色
                            for col_idx in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = row_fill
                        except:
                            pass
                    
                    # 格式化盈亏列的字体颜色
                    for col_name, col_idx in pnl_columns.items():
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        try:
                            value = float(cell.value) if cell.value is not None else 0
                            if value > 0:
                                cell.font = positive_font
                                cell.number_format = '#,##0.00'
                            elif value < 0:
                                cell.font = negative_font
                                cell.number_format = '#,##0.00'
                            cell.alignment = right_alignment
                        except:
                            pass
                
                # 格式化金额列
                money_cols = ['持仓市值', '持仓成本总额', '累计买入金额', '累计卖出金额']
                for col_name in money_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = right_alignment
                            cell.number_format = '#,##0.00'
                
                # 格式化价格列
                price_cols = ['持仓成本价', '当前价格', '收盘价', '平均买入价', '平均卖出价']
                for col_name in price_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = right_alignment
                            cell.number_format = '#,##0.0000'
                
                # 格式化日期列
                date_cols = ['日期', '首次交易日期', '最后交易日期']
                for col_name in date_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_alignment
                            cell.number_format = 'yyyy-mm-dd'
                            cell.font = date_font
            
            elif sheet_type == 'positions':
                # 持仓数据特定格式
                # 为每个单元格添加边框
                for row_idx in range(2, worksheet.max_row + 1):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                
                # 格式化盈亏列
                pnl_columns = {}
                for col_name, col_idx in column_indices.items():
                    if '盈亏' in col_name or '比例' in col_name:
                        pnl_columns[col_name] = col_idx
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            try:
                                value = float(cell.value) if cell.value is not None else 0
                                if value > 0:
                                    cell.font = positive_font
                                elif value < 0:
                                    cell.font = negative_font
                                cell.alignment = right_alignment
                                cell.number_format = '#,##0.00'
                            except:
                                pass
                
                # 格式化金额列
                money_cols = ['持仓市值', '持仓成本总额', '买入手续费', '卖出手续费', '总手续费']
                for col_name in money_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = right_alignment
                            cell.number_format = '#,##0.00'
                
                # 格式化价格列
                price_cols = ['持仓成本价', '当前价格', '平均买入价', '平均卖出价']
                for col_name in price_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = right_alignment
                            cell.number_format = '#,##0.0000'
                
                # 格式化日期列
                date_cols = ['首次交易日期', '最后交易日期']
                for col_name in date_cols:
                    if col_name in column_indices:
                        col_idx = column_indices[col_name]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_alignment
                            cell.number_format = 'yyyy-mm-dd'
                            cell.font = date_font
            
            elif sheet_type == 'dividends':
                # 分红记录特定格式
                # 找到金额相关列和日期列
                money_columns = {}
                date_col = None
                
                for col_name, col_idx in column_indices.items():
                    if '金额' in col_name or '分红' in col_name or '税费' in col_name:
                        money_columns[col_name] = col_idx
                    if col_name == '日期':
                        date_col = col_idx
                
                # 格式化数据行
                for row_idx in range(2, worksheet.max_row + 1):
                    # 为每个单元格添加边框和居中对齐
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = center_alignment
                    
                    # 格式化金额列
                    for col_name, col_idx in money_columns.items():
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        try:
                            value = float(cell.value) if cell.value is not None else 0
                            # 设置金额格式和字体
                            cell.font = money_font
                            cell.alignment = right_alignment
                            cell.number_format = '#,##0.00'
                        except:
                            pass
                    
                    # 格式化日期列
                    if date_col:
                        cell = worksheet.cell(row=row_idx, column=date_col)
                        cell.font = date_font
                        cell.alignment = center_alignment
                        # 设置日期格式
                        cell.number_format = 'yyyy-mm-dd'
                
                # 如果有数据行，添加合计行
                if worksheet.max_row > 1:
                    total_row = worksheet.max_row + 1
                    
                    # 添加合计标签
                    worksheet.cell(row=total_row, column=1).value = "合计"
                    worksheet.cell(row=total_row, column=1).font = Font(bold=True)
                    worksheet.cell(row=total_row, column=1).alignment = center_alignment
                    worksheet.cell(row=total_row, column=1).border = thin_border
                    
                    # 为金额列添加合计公式
                    for col_name, col_idx in money_columns.items():
                        cell = worksheet.cell(row=total_row, column=col_idx)
                        # 设置求和公式
                        start_cell = f"{get_column_letter(col_idx)}2"
                        end_cell = f"{get_column_letter(col_idx)}{total_row - 1}"
                        cell.value = f"=SUM({start_cell}:{end_cell})"
                        
                        # 设置格式
                        cell.font = Font(color='006600', bold=True)
                        cell.alignment = right_alignment
                        cell.number_format = '#,##0.00'
                        cell.border = thin_border
                    
                    # 为其他列添加空单元格并设置边框
                    for col_idx in range(2, worksheet.max_column + 1):
                        if col_idx not in money_columns.values():
                            cell = worksheet.cell(row=total_row, column=col_idx)
                            cell.border = thin_border
            
            # 自适应列宽调整
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # 检查标题行
                header_length = len(str(column[0].value)) * 2  # 标题行字体加粗，需要更多空间
                max_length = max(max_length, header_length)
                
                # 检查数据行
                for cell in column[1:]:
                    try:
                        cell_length = len(str(cell.value))
                        # 根据数据类型调整宽度
                        if isinstance(cell.value, (int, float)):
                            cell_length = cell_length * 1.5  # 数字需要更多空间
                        elif isinstance(cell.value, datetime):
                            cell_length = 12  # 日期固定宽度
                        max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # 设置列宽，最小8，最大30
                adjusted_width = max(8, min(max_length + 2, 30))
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 添加交替行颜色
            alternate_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')  # 浅灰色
            
            # 如果不是交易明细或盈亏分析（它们已经有自己的行颜色），则添加交替行颜色
            if sheet_type not in ['trades', 'pnl', 'stock_pnl']:
                for row_idx in range(2, worksheet.max_row + 1):
                    if row_idx % 2 == 0:  # 偶数行
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if not cell.fill.start_color.index or cell.fill.start_color.index == 'FFFFFFFF':
                                cell.fill = alternate_fill
            
            logger.info(f"工作表 '{sheet_name}' 格式化完成")
            
        except Exception as e:
            logger.warning(f"工作表 '{sheet_name}' 格式化失败: {e}")
    
    def save_results(self, output_file):
        """
            保存分析结果到单个Excel文件的不同工作表
            
            Args:
                output_file: 输出Excel文件路径
                
            Returns:
                是否成功保存结果
        """
        try:
            # 获取持仓数据
            positions_df = self.get_current_positions()
            
            # 检查是否有数据可保存
            has_pnl_data = self.daily_pnl is not None and not self.daily_pnl.empty
            has_trades_data = self.trades_df is not None and not self.trades_df.empty
            has_positions_data = not positions_df.empty
            
            if not (has_pnl_data or has_trades_data or has_positions_data):
                logger.warning("没有数据可保存")
                return False
            
            # 生成股票历史盈亏数据
            stock_pnl_df = self.get_stock_historical_pnl()
            has_stock_pnl_data = not stock_pnl_df.empty
            
            # 检查是否有证券信息数据
            has_securities_data = self.securities_df is not None and not self.securities_df.empty
            
            # 检查是否有分红记录数据
            has_dividend_data = self.dividend_df is not None and not self.dividend_df.empty
            
            # 使用ExcelWriter将多个DataFrame保存到不同工作表
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 保存盈亏分析
                if has_pnl_data:
                    # 按照时间倒序和持仓数量倒序排列
                    sorted_daily_pnl = self.daily_pnl.sort_values(['日期', '持仓数量'], ascending=[False, False])
                    sorted_daily_pnl.to_excel(writer, sheet_name='盈亏分析', index=False)
                    self._format_sheet(writer, '盈亏分析', sheet_type='pnl')
                    logger.info("盈亏分析已保存到工作表 '盈亏分析'")
                else:
                    logger.warning("没有盈亏数据可保存")
                
                # 保存交易明细（带颜色格式）
                if has_trades_data:
                    sorted_trades_df = self.trades_df.sort_values(['日期', '成交数量'], ascending=[False, False])
                    sorted_trades_df.to_excel(writer, sheet_name='交易明细', index=False)
                    self._format_sheet(writer, '交易明细', sheet_type='trades')
                    logger.info("交易明细已保存到工作表 '交易明细'")
                else:
                    logger.warning("没有交易数据可保存")
                
                # 保存持仓数据
                if has_positions_data:
                    positions_df.to_excel(writer, sheet_name='持仓数据', index=False)
                    self._format_sheet(writer, '持仓数据', sheet_type='positions')
                    logger.info("持仓数据已保存到工作表 '持仓数据'")
                else:
                    logger.warning("没有持仓数据可保存")
                
                # 保存股票历史盈亏
                if has_stock_pnl_data:
                    stock_pnl_df.to_excel(writer, sheet_name='股票历史盈亏', index=False)
                    self._format_sheet(writer, '股票历史盈亏', sheet_type='stock_pnl')
                    logger.info("股票历史盈亏已保存到工作表 '股票历史盈亏'")
                else:
                    logger.warning("没有股票历史盈亏数据可保存")
                
                # 保存分红记录
                sorted_dividends_df = self.dividend_df.sort_values('日期', ascending=False).reset_index(drop=True)
                sorted_dividends_df.to_excel(writer, sheet_name='分红记录', index=False)
                self._format_sheet(writer, '分红记录', sheet_type='dividends')
                logger.info("分红记录已保存到工作表 '分红记录'")
                
            
            logger.info(f"所有分析结果已保存到: {output_file}")
            return True
        except Exception as e:
            logger.error(f"保存结果失败: {e}")
            return False


def main():
    # 主函数
    # 设置根目录为当前目录下的 data 目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.join(current_dir, "data")
    
    # 确保目录存在
    if not os.path.exists(root_dir):
        os.makedirs(root_dir)
        logger.info(f"已创建目录: {root_dir}")
    
    # 输入文件路径
    input_file = os.path.join(root_dir, "交易数据.xlsx")
    
    # 输出文件路径
    output_file = os.path.join(root_dir, "交易分析结果.xlsx")
    
    logger.info(f"输入文件路径: {input_file}")
    logger.info(f"输出文件路径: {output_file}")
    
    # 创建交易数据处理器
    processor = TradingProcessor()
    
    # 加载数据
    logger.info("正在加载数据...")
    if not processor.load_data(input_file, trades_sheet='交易数据', rates_sheet='费率配置', prices_sheet='收盘价格'):
        logger.error("数据加载失败，程序终止")
        return
    
    # 处理数据
    logger.info("正在处理数据...")
    if not processor.process_data():
        logger.error("数据处理失败，程序终止")
        return
    
    # 保存结果
    logger.info("正在保存结果...")
    if not processor.save_results(output_file):
        logger.error("结果保存失败")
        return
    
    logger.info("处理完成！")
    logger.info(f"分析结果已保存到: {output_file}")


if __name__ == "__main__":
    main()